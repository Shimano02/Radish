import os
import uuid
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

import httpx
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Load environment vars from .env
load_dotenv()

# Logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Env vars
DIFY_API_URL = os.getenv("DIFY_API_URL", "https://api.dify.ai/v1").rstrip("/")
DIFY_API_KEY = os.getenv("DIFY_API_KEY", "").strip()

# Validate
if not DIFY_API_KEY:
    raise RuntimeError("Missing DIFY_API_KEY in environment")

# Excel configuration
EXCEL_LOG_DIR = Path("interview_logs")
EXCEL_LOG_DIR.mkdir(exist_ok=True)

# FastAPI setup
app = FastAPI(title="AI Interviewer API")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], 
    allow_credentials=True, 
    allow_methods=["*"], 
    allow_headers=["*"]
)

# Models
class InterviewerInfo(BaseModel):
    id: str
    name: str
    title: str
    description: str
    image_url: str

class ChatMessage(BaseModel):
    message: str
    interviewer_id: str
    conversation_id: str

class ChatResponse(BaseModel):
    response: str
    conversation_id: str
    stage: int

# Excel utilities
class ExcelLogger:
    """Excel面接ログ管理クラス"""
    
    @staticmethod
    def get_excel_file_path() -> Path:
        """日付ベースのExcelファイルパスを取得"""
        date_str = datetime.now().strftime('%Y-%m-%d')
        return EXCEL_LOG_DIR / f"interview_log_{date_str}.xlsx"
    
    @staticmethod
    def create_workbook_with_headers() -> Workbook:
        """ヘッダー付きの新規Excelワークブックを作成"""
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "面接ログ"
        
        # ヘッダー設定
        headers = [
            "タイムスタンプ", "会話ID", "面接官ID", "面接官名", 
            "ユーザーメッセージ", "AI応答", "対話回数", "ステージ"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="366092", 
                end_color="366092", 
                fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 列幅調整
        column_widths = [20, 15, 15, 10, 50, 50, 10, 10]
        for col_num, width in enumerate(column_widths, 1):
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].width = width
        
        return workbook
    
    @classmethod
    async def write_interview_log(
        cls,
        conversation_id: str,
        interviewer_id: str,
        interviewer_name: str,
        user_message: str,
        ai_response: str,
        dialogue_count: int,
        stage: int
    ) -> bool:
        """面接ログをExcelファイルに書き込み"""
        try:
            excel_file = cls.get_excel_file_path()
            
            # Excelファイルの存在確認と読み込み
            if excel_file.exists():
                workbook = load_workbook(excel_file)
                worksheet = workbook.active
            else:
                workbook = cls.create_workbook_with_headers()
                worksheet = workbook.active
            
            # 新しい行を追加
            new_row = worksheet.max_row + 1
            timestamp = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')
            
            # データを書き込み
            row_data = [
                timestamp,
                conversation_id,
                interviewer_id,
                interviewer_name,
                user_message,
                ai_response,
                dialogue_count,
                stage
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=new_row, column=col_num)
                cell.value = value
                
                # セルの書式設定
                if col_num in [1, 2, 3, 4, 7, 8]:  # タイムスタンプ、ID、数値列
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:  # テキスト列
                    cell.alignment = Alignment(
                        horizontal="left", 
                        vertical="top", 
                        wrap_text=True
                    )
            
            # ファイルを保存
            workbook.save(excel_file)
            workbook.close()
            
            logger.info(f"面接ログをExcelファイルに保存: {excel_file}")
            return True
            
        except Exception as e:
            logger.error(f"Excelログ保存エラー: {e}")
            return False

# Dify API utilities
class DifyAPIClient:
    """Dify API呼び出し管理クラス"""
    
    def __init__(self, api_url: str, api_key: str):
        self.api_url = api_url
        self.api_key = api_key
    
    async def send_chat_message(
        self,
        message: str,
        interviewer_id: str,
        conversation_id: str,
        dialogue_count: int,
        dify_conversation_id: Optional[str] = None
    ) -> Dict[str, Any]:
        """Dify APIにチャットメッセージを送信"""
        payload = {
            "inputs": {
                "user_message": message,
                "interviewer_type": interviewer_id,
                "dialogue_count": dialogue_count
            },
            "query": message,
            "user": conversation_id,
            "stream": False
        }
        
        if dify_conversation_id:
            payload["conversation_id"] = dify_conversation_id
        
        logger.info(f"Calling Dify API with payload: {payload}")
        
        try:
            async with httpx.AsyncClient(timeout=30) as client:
                response = await client.post(
                    f"{self.api_url}/chat-messages",
                    headers={
                        "Authorization": f"Bearer {self.api_key}",
                        "Content-Type": "application/json"
                    },
                    json=payload
                )
                response.raise_for_status()
                return response.json()
                
        except httpx.HTTPStatusError as e:
            logger.error(f"Dify API error {e.response.status_code}: {e.response.text}")
            raise HTTPException(status_code=502, detail="Error from Dify API")
        except Exception as e:
            logger.error(f"Unexpected Dify API error: {e}")
            raise HTTPException(status_code=500, detail="Dify API call failed")

# Session management
class SessionManager:
    """セッション管理クラス"""
    
    def __init__(self):
        self.conversation_sessions: Dict[str, Any] = {}
    
    def create_session(self, interviewer_id: str) -> str:
        """新しいセッションを作成"""
        session_id = str(uuid.uuid4())
        self.conversation_sessions[session_id] = {
            "interviewer_id": interviewer_id,
            "stage": 0,
            "dialogue_count": 0,
            "dify_id": None,
            "messages": [],
            "created_at": datetime.now(timezone.utc)
        }
        return session_id
    
    def get_session(self, session_id: str) -> Optional[Dict[str, Any]]:
        """セッションを取得"""
        return self.conversation_sessions.get(session_id)
    
    def update_session(
        self,
        session_id: str,
        user_message: str,
        ai_response: str,
        dify_conversation_id: Optional[str] = None
    ) -> None:
        """セッションを更新"""
        session = self.conversation_sessions.get(session_id)
        if not session:
            return
        
        session["dialogue_count"] += 1
        session["stage"] += 1
        
        if dify_conversation_id and not session.get("dify_id"):
            session["dify_id"] = dify_conversation_id
        
        timestamp = datetime.now(timezone.utc).isoformat()
        session["messages"].append({
            "user": user_message,
            "ai": ai_response,
            "timestamp": timestamp
        })

# In-memory data
INTERVIEWERS: List[Dict[str, Any]] = [
    {
        "id": "construction_engineer",
        "name": "不動",
        "title": "施設長",
        "description": "１次面接を担当します。",
        #"image_url": "/assets/construction_engineer.png"
        "image_url": "/assets/dayCare.png"
    }

]

# Initialize services
dify_client = DifyAPIClient(DIFY_API_URL, DIFY_API_KEY)
session_manager = SessionManager()

# Helper functions
def get_interviewer_by_id(interviewer_id: str) -> Optional[Dict[str, Any]]:
    """IDから面接官情報を取得"""
    return next((i for i in INTERVIEWERS if i["id"] == interviewer_id), None)

def extract_ai_response(dify_data: Dict[str, Any]) -> str:
    """Dify APIレスポンスからAI応答を抽出"""
    return (
        dify_data.get("answer")
        or dify_data.get("choices", [{}])[0].get("message", {}).get("content", "")
        or "申し訳ございませんが、応答を取得できませんでした。"
    )

# API Endpoints
@app.get("/api/interviewers", response_model=List[InterviewerInfo])
async def get_interviewers():
    """面接官一覧を取得"""
    return INTERVIEWERS

@app.post("/api/chat/start")
async def start_chat(interviewer_id: str):
    """面接チャットを開始"""
    interviewer = get_interviewer_by_id(interviewer_id)
    if not interviewer:
        raise HTTPException(status_code=400, detail="Invalid interviewer ID")
    
    session_id = session_manager.create_session(interviewer_id)
    greeting = "準備ができましたら『開始』を押して下さい。音声でも入力可能です。"
    
    logger.info(f"Started new interview session: {session_id} with {interviewer_id}")
    
    return {
        "conversation_id": session_id,
        "greeting": greeting,
        "interviewer": interviewer
    }

@app.post("/api/chat/message", response_model=ChatResponse)
async def send_message(chat_message: ChatMessage):
    """メッセージを送信し、AI応答を取得"""
    # 1) セッション確認
    session = session_manager.get_session(chat_message.conversation_id)
    if not session:
        raise HTTPException(status_code=400, detail="Invalid conversation ID")
    
    # 2) 面接官情報取得
    interviewer = get_interviewer_by_id(chat_message.interviewer_id)
    if not interviewer:
        raise HTTPException(status_code=400, detail="Invalid interviewer ID")
    
    # 3) Dify API 呼び出し
    dify_response = await dify_client.send_chat_message(
        message=chat_message.message,
        interviewer_id=chat_message.interviewer_id,
        conversation_id=chat_message.conversation_id,
        dialogue_count=session["dialogue_count"] + 1,
        dify_conversation_id=session.get("dify_id")
    )
    
    # 4) AI応答抽出
    ai_response = extract_ai_response(dify_response)
    
    # 5) セッション更新
    session_manager.update_session(
        session_id=chat_message.conversation_id,
        user_message=chat_message.message,
        ai_response=ai_response,
        dify_conversation_id=dify_response.get("conversation_id")
    )
    
    # 6) 更新されたセッション情報を取得
    updated_session = session_manager.get_session(chat_message.conversation_id)
    
    # 7) Excelログに書き込み
    await ExcelLogger.write_interview_log(
        conversation_id=chat_message.conversation_id,
        interviewer_id=chat_message.interviewer_id,
        interviewer_name=interviewer["name"],
        user_message=chat_message.message,
        ai_response=ai_response,
        dialogue_count=updated_session["dialogue_count"],
        stage=updated_session["stage"]
    )
    
    # 8) レスポンス返却
    return ChatResponse(
        response=ai_response,
        conversation_id=chat_message.conversation_id,
        stage=updated_session["stage"]
    )

@app.get("/health")
async def health_check():
    """ヘルスチェック"""
    return {
        "status": "healthy",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "excel_log_dir": str(EXCEL_LOG_DIR.absolute())
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        app, 
        host="0.0.0.0",
        port=int(os.getenv("PORT", 8000)), 
        reload=True
    )
